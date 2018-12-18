#!/usr/bin/python3
# coding:utf-8
# @author : gswu
# @date :2018/11/28 14:04

# 读写xls表格03-07
import xlrd
import xlwt
#读写xlsx表格07版
import openpyxl
from tkinter import font

def write_03_excel(myfilepath):
    #myfilepath="F:/testpy.xls"
    #初始化一个excel
    wb = xlwt.Workbook(encoding='utf-8')
    #新建一个sheet
    sheet = wb.add_sheet("test01") #添加表格名称
    mystyle=xlwt.XFStyle() #初始化样式
    font=xlwt.Font()#创建字体
    font.name=u'微软雅黑' #字体类型
    font.colour_index=6 #字体颜色
    font.underline = True #下划线
    font.italic=True#斜体
    font.height=200#字体大小，200等于excel字体大小中的10
    mystyle.font=font #设定样式
    value = [["姓名", "年龄", "电话", "婚姻状况","备注"],
             ["范彬彬", "22", "18888888888", "已婚","学生"],
             ["袁姗姗", "25", "18999999999", "未婚","学生"],
             ["刘德华", "50", "17777777777", "已婚","团支书"],
             ["张学友", "55", "15555555555", "已婚","学生"],
             ["郭富城", "55", "13333333333", "已婚","班长"]]
    for i in range(0,len(value)):
        for e in range(0,len(value[i])):
            sheet.write(i,e,value[i][e],mystyle)  #i和e分别表示行和列
    
    wb.save(myfilepath)
    print("写入表格成功","路径为:",myfilepath)
    

def read_03_excel(myfilepath):    
    #myfilepath="F:/testpy.xls"
    wb=xlrd.open_workbook(myfilepath)#打开excel文件
    r_sheet=wb.sheet_names()#查找所有表的名字
    print("表名字：",r_sheet)
    work_sheet=wb.sheet_by_name(r_sheet[0])
    print("work_sheet：",work_sheet)
    for i in range(0,work_sheet.nrows): #循环所有行
        row=work_sheet.row(i)  #获取第i行
        print("第 %d行:" %(i),row)
        for j in range(0,work_sheet.ncols):
            print("第%2d行，第%5d列单元格为："%(i,j),work_sheet.cell_value(i,j))

#read_03_excel("F:/testdemo.xlsx")    

#myfilepath="F:/test07py.xlsx"
def wite_07_excel(myfilepath):
    wb=openpyxl.Workbook() #打开文件
    mysheet=wb.create_sheet(title="第一个sheet页", index=0)
    #print(wb.get_sheet_names())
    #mysheet.title ="test_sheet" #添加sheet页标题
    myvalue = [["姓名", "年龄", "电话", "婚姻状况"],
             ["范彬彬", "22", "18888888888", "已婚"],
             ["袁姗姗", "25", "18999999999", "未婚"],
             ["刘德华", "50", "17777777777", "已婚"],
             ["张学友", "55", "15555555555", "已婚"],
             ["郭富城", "55", "13333333333", "已婚"]]
    for i in range(0,len(myvalue)):
        for j in range(0,len(myvalue[i])):
            mysheet.cell(row=i+1, column=j+1, value=myvalue[i][j])#写入单元格
    wb.save(myfilepath)
    print("07表格写入成功")
    
